import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def check_passed_rows():
    passed_results = []

    URL = input("Enter URL: ")
    driver = webdriver.Chrome()
    driver.get(URL)
    WebDriverWait(driver, 2).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id='suite-stats']/tbody/tr[1]/td[1]"))
    )

    first_row_links = driver.find_elements(By.XPATH, "//*[@id='suite-stats']/tbody/tr[1]/td[1]")

    for first_row in first_row_links:
        first_row.click()
        
        columns = driver.find_elements(By.XPATH, "//*[@id='test-details']/tbody/tr/td[1]")
        status = driver.find_elements(By.XPATH, "//*[@id='test-details']/tbody/tr/td[5]")

        for i in range(len(columns)):
            col_text = columns[i].text
            stat_text = status[i].text
            
            col_parts = col_text.split(":")
            col_value = col_parts[1] if len(col_parts) > 1 else col_text
            
            if stat_text.lower() == "pass":
                passed_results.append((col_value, stat_text))
    
    driver.quit()
    save_passed_results(passed_results)

def save_passed_results(passed_results):
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "TestResults")
    if not os.path.exists(desktop_path):
        os.makedirs(desktop_path)
    
    file_name = f"PassedResults_{int(time.time())}.xlsx"
    file_path = os.path.join(desktop_path, file_name)
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Passed Tests"
    
    sheet.append(["Test Name", "Status"])
    
    fill_color = PatternFill(start_color='00b050', end_color='00b050', fill_type='solid')
    
    for result in passed_results:
        sheet.append(result)
        row_num = sheet.max_row  # Get the last row number
        status_cell = sheet.cell(row=row_num, column=2)  # Apply color only to the status column
        status_cell.fill = fill_color
    import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def check_passed_rows():
    passed_results = []

    URL = input("Enter URL: ")
    driver = webdriver.Chrome()
    driver.get(URL)
    WebDriverWait(driver, 2).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id='suite-stats']/tbody/tr[1]/td[1]"))
    )

    first_row_links = driver.find_elements(By.XPATH, "//*[@id='suite-stats']/tbody/tr[1]/td[1]")

    for first_row in first_row_links:
        first_row.click()
        
        columns = driver.find_elements(By.XPATH, "//*[@id='test-details']/tbody/tr/td[1]")
        status = driver.find_elements(By.XPATH, "//*[@id='test-details']/tbody/tr/td[5]")

        for i in range(len(columns)):
            col_text = columns[i].text
            stat_text = status[i].text
            
            col_parts = col_text.split(":")
            col_value = col_parts[1] if len(col_parts) > 1 else col_text
            
            if stat_text.lower() == "pass":
                passed_results.append((col_value, stat_text))
    
    driver.quit()
    save_passed_results(passed_results)

def save_passed_results(passed_results):
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "TestResults")
    if not os.path.exists(desktop_path):
        os.makedirs(desktop_path)
    
    file_name = f"PassedResults_{int(time.time())}.xlsx"
    file_path = os.path.join(desktop_path, file_name)
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Passed Tests"
    
    sheet.append(["Test Name", "Status"])
    
    fill_color = PatternFill(start_color='00b050', end_color='00b050', fill_type='solid')
    
    for result in passed_results:
        sheet.append(result)
        row_num = sheet.max_row  # Get the last row number
        status_cell = sheet.cell(row=row_num, column=2)  # Apply color only to the status column
        status_cell.fill = fill_color

    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[col_letter].width = max_length + 2  # Add padding  
    workbook.save(file_path)
    print(f"Results saved at: {file_path}")

if __name__ == "__main__":
    check_passed_rows()
