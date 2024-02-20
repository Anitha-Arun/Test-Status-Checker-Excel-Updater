from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from fuzzywuzzy import fuzz


def check_failed_rows():
    result_list = []

    URL = input("Enter URL: ")
    BROWSER = "chrome"

    driver = webdriver.Chrome()
    driver.get(URL)
    WebDriverWait(driver, 2).until(
        EC.presence_of_element_located((By.XPATH, "//*[@id='suite-stats']/tbody/tr[1]/td[1]")))

    first_row_links = driver.find_elements(By.XPATH, "//*[@id='suite-stats']/tbody/tr[1]/td[1]")

    for first_row in first_row_links:
        first_row.click()

        columns = driver.find_elements(By.XPATH, "//*[@id='test-details']/tbody/tr/td[1]")
        status = driver.find_elements(By.XPATH, "//*[@id='test-details']/tbody/tr/td[5]")

        for i in range(len(columns)):
            col_text = columns[i].text
            stat_text = status[i].text

            col_parts = col_text.split(":")
            col_value = col_parts[1] if len(col_parts) > 1 else col_text

            result_list.append(col_value)
            result_list.append(stat_text)
            print(f"{col_value} {stat_text}")

    driver.quit()
# excel sheet
    FILE_PATH = input("Enter file path: ")
    SHEET_NAME = input("Enter sheet name: ")
    update_excel_sheet(FILE_PATH, SHEET_NAME, result_list)


def update_excel_sheet(file_path, sheet_name, result_list):
    workbook = load_workbook(file_path)

    if sheet_name not in workbook.sheetnames:
        print(f"Sheet '{sheet_name}' not found in the workbook.")
        workbook.close()
        return

    sheet = workbook[sheet_name]
    matched_rows = []  # Move this line to the beginning of the function

    for row_index in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_index, column=2).value

        if cell_value is not None:
            match_found = False

            for value in result_list:
                similarity_score = fuzz.ratio(value.lower(), cell_value.lower())

                if similarity_score >= 80:
                    match_found = True

                    next_index = (result_list.index(value) + 1) % len(result_list)
                    next_value = result_list[next_index]

                    sheet.cell(row=row_index, column=3, value=next_value)

                    fill_color = '00b050' if next_value.lower() == 'pass' else 'FF0000'
                    sheet.cell(row=row_index, column=3).fill = PatternFill(start_color=fill_color, end_color=fill_color,
                                                                           fill_type='solid')

                    matched_rows.append(f"Match found in row {row_index}: {value} -> {next_value}")

                    break



    workbook.save(file_path)
    workbook.close()

    if matched_rows:
        print("\nMatched Rows:")
        for row_detail in matched_rows:
            print(row_detail)


if __name__ == "__main__":
    check_failed_rows()